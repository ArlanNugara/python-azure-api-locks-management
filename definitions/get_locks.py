import requests
import sys
import json
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from datetime import datetime
import time

# Get today's date
todays_date = datetime.today().strftime('%Y-%m-%d')

# Declare the CSV header here
subscription_header = ['Subscription Name', 'Subscription ID', 'Scope', 'Lock Scope', 'Lock Name', 'Lock Level', 'Lock Notes', 'Create Locks', 'Update Locks', 'Delete Locks']
rg_header = ['Subscription Name', 'Subscription ID', 'Scope', 'Resource Group Name', 'Location', 'Lock Scope', 'Lock Name', 'Lock Level', 'Lock Notes', 'Create Locks', 'Update Locks', 'Delete Locks']
resource_header = ['Subscription Name', 'Subscription ID', 'Scope', 'Resource Group Name', 'Location', 'Resource Name', 'Resource Type', 'Lock Scope', 'Lock Name', 'Lock Level', 'Lock Notes', 'Create Locks', 'Update Locks', 'Delete Locks']

# Start API calls
def get_all_locks(main_url, subscription_id, header):
    # Get Subscription Details
    print("Getting Details for Subscription ID - "+subscription_id+"")
    get_subscription_details = requests.get(url = ""+main_url+"/subscriptions/"+subscription_id+"?api-version=2020-01-01", headers = header)
    subscription_response_to_json = get_subscription_details.json()
    subscription_lock_details = requests.get(url = ""+main_url+""+subscription_response_to_json["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
    subscription_lock_response_to_json = subscription_lock_details.json()
    if subscription_lock_details.status_code == 200 or subscription_lock_details.status_code == 204:
        print('Recieved API response for Locks')
        lock_excel_value = []
        if 'value' in subscription_lock_response_to_json:
            for locks in subscription_lock_response_to_json['value']:
                lock_level = locks['properties']['level']
                if 'notes' in locks['properties']:
                    lock_notes = locks['properties']['notes']
                else:
                    lock_notes = None
                scope_split = locks['id'].split('/')[:-4]
                scope_joined = '/'.join(scope_split)
                lock_id = locks['id']
                lock_name = locks['name']
                subscription_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], scope_joined, lock_id, lock_name, lock_level, lock_notes, None, None, None]
                lock_excel_value.append(subscription_lock_excel_values)
        else:
            print('No locks found for ', subscription_response_to_json['displayName'])
            subscription_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], None, None, None, None, None, None, None, None]
            lock_excel_value.append(subscription_lock_excel_values)
    else:
        print('Unable to get Locks info from Azure API. Please try again later')
        sys.exit(1)
    subscription_excel_header = pd.DataFrame(columns = subscription_header)
    subscription_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl')
    subscription_excel_header.to_excel(subscription_excel_writer, sheet_name='Subscription', index=False)
    subscription_excel_writer.save()
    print("Successfuly created the Subscription Worksheet")
    subscription_workbook = load_workbook(filename = "Reports.xlsx")
    subscription_worksheet = subscription_workbook["Subscription"]
    for subscription_rows in dataframe_to_rows(pd.DataFrame(lock_excel_value), index=False, header=False):
        subscription_worksheet.append(subscription_rows)
    subscription_worksheet.auto_filter.ref = subscription_worksheet.dimensions
    subscription_workbook.save("Reports.xlsx")
    subscription_workbook.close()
    print("Finished writing Subscription Details to Excel File")

    # Get list of resource groups
    print("Getting Resource Groups Details for Subscription ID - "+subscription_id+"")
    get_rg_details = requests.get(url = ""+main_url+"/subscriptions/"+subscription_id+"/resourcegroups?api-version=2021-04-01", headers = header)
    rg_response_to_json = get_rg_details.json()
    rg_values = rg_response_to_json["value"]
    resource_group_excel_header = pd.DataFrame(columns = rg_header)
    resource_group_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl', mode='a')
    resource_group_excel_header.to_excel(resource_group_excel_writer, sheet_name='ResourceGroup', index=False)
    resource_group_excel_writer.save()
    print("Successfuly created the Resource Group Worksheet")
    rg_lock_data = []
    for rg in rg_values:
        rg_lock_details = requests.get(url = ""+main_url+""+rg["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
        rg_lock_resonse_to_json = rg_lock_details.json()
        if rg_lock_details.status_code == 200 or rg_lock_details.status_code == 204:
            rg_lock_response_not_null = rg_lock_resonse_to_json["value"]
            if rg_lock_response_not_null:
                for locks in rg_lock_response_not_null:
                    if 'notes' in locks['properties']:
                        lock_notes = locks['properties']['notes']
                    else:
                        lock_notes = None
                    lock_level = locks['properties']['level']
                    scope_split = locks['id'].split('/')[:-4]
                    scope_joined = '/'.join(scope_split)
                    lock_id = locks['id']
                    lock_name = locks['name']
                    rg_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], scope_joined, rg["name"], rg["location"], lock_id, lock_name, lock_level, lock_notes, None, None, None]
            else:
                rg_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], None, rg["name"], rg["location"], None, None, None, None, None, None, None]
        else:
            rg_lock_error_message = rg_lock_resonse_to_json['error']['message']
            rg_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], None, rg["name"], rg["location"], None, None, None, json.dumps(rg_lock_error_message, indent=2), None, None, None]
        rg_lock_data.append(rg_lock_excel_values)
    print("Resource Group API Response Received for Subscription ID - "+subscription_id+"")
    resource_group_workbook = load_workbook(filename = "Reports.xlsx")
    resource_group_worksheet = resource_group_workbook["ResourceGroup"]
    for resource_group_rows in dataframe_to_rows(pd.DataFrame(rg_lock_data), index=False, header=False):
        resource_group_worksheet.append(resource_group_rows)
    resource_group_worksheet.auto_filter.ref = resource_group_worksheet.dimensions
    resource_group_workbook.save("Reports.xlsx")
    resource_group_workbook.close()
    print("Finished writing Resource Group Details to Excel File")
    
    # Get list of resources
    resource_excel_header = pd.DataFrame(columns = resource_header)
    resource_excel_writer = pd.ExcelWriter('Reports.xlsx', engine='openpyxl', mode='a')
    resource_excel_header.to_excel(resource_excel_writer, sheet_name='Resource', index=False)
    resource_excel_writer.save()
    print("Successfully created the Resource Worksheet")
    for rg in rg_values:
        print("Getting Resources Details for Resource Group - "+rg["name"]+"")
        get_resource_details = requests.get(url = ""+main_url+""+rg["id"]+"/resources?api-version=2021-04-01", headers = header)
        resource_response_to_json = get_resource_details.json()
        print("Resource API Response Received for Resource Group - "+rg["name"]+"")
        resource_values = resource_response_to_json["value"]
        resource_lock_data = []
        for resource in resource_values:
            resource_lock_details = requests.get(url = ""+main_url+""+resource["id"]+"/providers/Microsoft.Authorization/locks?api-version=2016-09-01", headers = header)
            resource_lock_resonse_to_json = resource_lock_details.json()
            if resource_lock_details.status_code == 200 or resource_lock_details.status_code == 204:
                resource_lock_response_not_null = resource_lock_resonse_to_json["value"]
                if resource_lock_response_not_null:
                    for locks in resource_lock_response_not_null:
                        if 'notes' in locks['properties']:
                            lock_notes = locks['properties']['notes']
                        else:
                            lock_notes = None
                    lock_level = locks['properties']['level']
                    scope_split = locks['id'].split('/')[:-4]
                    scope_joined = '/'.join(scope_split)
                    lock_id = locks['id']
                    lock_name = locks['name']
                    resource_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], scope_joined, rg["name"], rg["location"], resource["name"], resource["type"], lock_id, lock_name, lock_level, lock_notes, None, None, None]
                else:
                    resource_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], None, rg["name"], rg["location"], resource["name"], resource["type"], None, None, None, None, None, None, None]
            else:
                resource_lock_error_message = resource_lock_resonse_to_json['error']['message']
                resource_lock_excel_values = [subscription_response_to_json['displayName'], subscription_response_to_json['subscriptionId'], None, rg["name"], rg["location"], resource["name"], resource["type"], None, None, None, json.dumps(resource_lock_error_message, indent=2), None, None, None]
            resource_lock_data.append(resource_lock_excel_values)
        resource_workbook = load_workbook(filename = "Reports.xlsx")
        resource_worksheet = resource_workbook["Resource"]
        for resource_rows in dataframe_to_rows(pd.DataFrame(resource_lock_data), index=False, header=False):
            resource_worksheet.append(resource_rows)
        resource_worksheet.auto_filter.ref = resource_worksheet.dimensions
        resource_workbook.save("Reports.xlsx")
        resource_workbook.close()
        time.sleep(1)
    print("Finished writing Resource Details to Excel File")